from flask import Blueprint, send_file, request, render_template, redirect, url_for, flash, jsonify, session
from database.models import db, Income, Expense, Budget, SavingsGoal, Settings
from datetime import datetime, date
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

export_bp = Blueprint('export', __name__)

def get_current_settings():
    user_id = session.get('user_id')
    settings = Settings.query.filter_by(user_id=user_id).first()
    if not settings:
        settings = Settings(currency='USD', theme='light', export_preference='excel', user_id=user_id)
        db.session.add(settings)
        db.session.commit()
    return settings

# Financial score helper for Power BI export (user specific)
def calculate_historical_score(user_id, year, month):
    # Fetch metrics for that month
    inc_sum = db.session.query(db.func.sum(Income.amount)).filter(
        Income.user_id == user_id,
        db.extract('year', Income.date) == year,
        db.extract('month', Income.date) == month
    ).scalar() or 0.0
    exp_sum = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.user_id == user_id,
        db.extract('year', Expense.date) == year,
        db.extract('month', Expense.date) == month
    ).scalar() or 0.0
    
    # Simple score based on savings rate for the month
    savings = max(0.0, inc_sum - exp_sum)
    savings_rate = (savings / inc_sum) * 100 if inc_sum > 0 else 0.0
    
    if savings_rate >= 30:
        return 90
    elif savings_rate >= 20:
        return 75
    elif savings_rate >= 10:
        return 60
    elif savings_rate > 0:
        return 45
    else:
        return 30

@export_bp.route('/export/excel')
def export_excel():
    settings = get_current_settings()
    user_id = session['user_id']
    
    # 1. Fetch Data for user
    incomes = Income.query.filter_by(user_id=user_id).order_by(Income.date.desc()).all()
    expenses = Expense.query.filter_by(user_id=user_id).order_by(Expense.date.desc()).all()
    budgets = Budget.query.filter_by(user_id=user_id).order_by(Budget.month.desc()).all()
    savings = SavingsGoal.query.filter_by(user_id=user_id).all()
    
    # Create Unified Transactions
    tx_list = []
    for inc in incomes:
        tx_list.append({
            'Type': 'Income',
            'Date': inc.date,
            'Category/Source': inc.source,
            'Amount': inc.amount,
            'Description': inc.description or ''
        })
    for exp in expenses:
        tx_list.append({
            'Type': 'Expense',
            'Date': exp.date,
            'Category/Source': exp.category,
            'Amount': exp.amount,
            'Description': exp.description or ''
        })
    tx_list.sort(key=lambda x: x['Date'], reverse=True)
    
    # 2. Setup Excel Writer
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Styling variables
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Neon Green
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="111827")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    total_border = Border(
        top=Side(style='thin', color='111827'),
        bottom=Side(style='double', color='111827')
    )
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # --- Sheet 1: Dashboard Summary ---
    ws1 = wb.active
    ws1.title = "Dashboard Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1['A1'] = "Student Finance Dashboard Summary"
    ws1['A1'].font = title_font
    ws1.row_dimensions[1].height = 30
    
    ws1['A3'] = "Metric"
    ws1['B3'] = "Value"
    for col in ['A3', 'B3']:
        ws1[col].fill = header_fill
        ws1[col].font = header_font
        ws1[col].alignment = align_center
    ws1.row_dimensions[3].height = 24
    
    total_inc = sum(i.amount for i in incomes)
    total_exp = sum(e.amount for e in expenses)
    balance = total_inc - total_exp
    total_saved = sum(s.current_amount for s in savings)
    
    metrics = [
        ("Total Income", total_inc),
        ("Total Expenses", total_exp),
        ("Current Balance", balance),
        ("Total Accumulated Savings", total_saved)
    ]
    
    for idx, (m_name, m_val) in enumerate(metrics, start=4):
        ws1.cell(row=idx, column=1, value=m_name).font = data_font
        ws1.cell(row=idx, column=1).border = thin_border
        
        val_cell = ws1.cell(row=idx, column=2, value=m_val)
        val_cell.font = data_font
        val_cell.number_format = f'$#,##0.00' if settings.currency == 'USD' else f'[$€-2] #,##0.00'
        val_cell.border = thin_border
        ws1.row_dimensions[idx].height = 20
        
    # --- Sheet 2: Income Data ---
    ws2 = wb.create_sheet(title="Income Data")
    ws2.views.sheetView[0].showGridLines = True
    ws2.append(["Date", "Source", "Amount", "Description"])
    
    for row in ws2[1]:
        row.fill = header_fill
        row.font = header_font
        row.alignment = align_center
    ws2.row_dimensions[1].height = 24
    
    for inc in incomes:
        ws2.append([inc.date, inc.source, inc.amount, inc.description or ''])
        
    for r in range(2, len(incomes) + 2):
        ws2.cell(row=r, column=1).alignment = align_center
        ws2.cell(row=r, column=1).number_format = 'YYYY-MM-DD'
        ws2.cell(row=r, column=3).number_format = '$#,##0.00'
        for c in range(1, 5):
            ws2.cell(row=r, column=c).font = data_font
            ws2.cell(row=r, column=c).border = thin_border
            
    # --- Sheet 3: Expense Data ---
    ws3 = wb.create_sheet(title="Expense Data")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["Date", "Category", "Amount", "Description"])
    
    for row in ws3[1]:
        row.fill = header_fill
        row.font = header_font
        row.alignment = align_center
    ws3.row_dimensions[1].height = 24
    
    for exp in expenses:
        ws3.append([exp.date, exp.category, exp.amount, exp.description or ''])
        
    for r in range(2, len(expenses) + 2):
        ws3.cell(row=r, column=1).alignment = align_center
        ws3.cell(row=r, column=1).number_format = 'YYYY-MM-DD'
        ws3.cell(row=r, column=3).number_format = '$#,##0.00'
        for c in range(1, 5):
            ws3.cell(row=r, column=c).font = data_font
            ws3.cell(row=r, column=c).border = thin_border
            
    # --- Sheet 4: Budget Data ---
    ws4 = wb.create_sheet(title="Budget Data")
    ws4.views.sheetView[0].showGridLines = True
    ws4.append(["Category", "Limit Amount", "Month"])
    
    for row in ws4[1]:
        row.fill = header_fill
        row.font = header_font
        row.alignment = align_center
    ws4.row_dimensions[1].height = 24
    
    for b in budgets:
        ws4.append([b.category, b.limit_amount, b.month])
        
    for r in range(2, len(budgets) + 2):
        ws4.cell(row=r, column=2).number_format = '$#,##0.00'
        ws4.cell(row=r, column=3).alignment = align_center
        for c in range(1, 4):
            ws4.cell(row=r, column=c).font = data_font
            ws4.cell(row=r, column=c).border = thin_border
            
    # --- Sheet 5: Savings Goals ---
    ws5 = wb.create_sheet(title="Savings Goals")
    ws5.views.sheetView[0].showGridLines = True
    ws5.append(["Goal Name", "Target Amount", "Current Amount", "Expected Completion"])
    
    for row in ws5[1]:
        row.fill = header_fill
        row.font = header_font
        row.alignment = align_center
    ws5.row_dimensions[1].height = 24
    
    for s in savings:
        ws5.append([s.name, s.target_amount, s.current_amount, s.target_date])
        
    for r in range(2, len(savings) + 2):
        ws5.cell(row=r, column=2).number_format = '$#,##0.00'
        ws5.cell(row=r, column=3).number_format = '$#,##0.00'
        ws5.cell(row=r, column=4).alignment = align_center
        ws5.cell(row=r, column=4).number_format = 'YYYY-MM-DD'
        for c in range(1, 5):
            ws5.cell(row=r, column=c).font = data_font
            ws5.cell(row=r, column=c).border = thin_border
            
    # --- Sheet 6: Transactions (Unified) ---
    ws6 = wb.create_sheet(title="Transactions")
    ws6.views.sheetView[0].showGridLines = True
    ws6.append(["Type", "Date", "Category/Source", "Amount", "Description"])
    
    for row in ws6[1]:
        row.fill = header_fill
        row.font = header_font
        row.alignment = align_center
    ws6.row_dimensions[1].height = 24
    
    for tx in tx_list:
        ws6.append([tx['Type'], tx['Date'], tx['Category/Source'], tx['Amount'], tx['Description']])
        
    for r in range(2, len(tx_list) + 2):
        ws6.cell(row=r, column=2).alignment = align_center
        ws6.cell(row=r, column=2).number_format = 'YYYY-MM-DD'
        ws6.cell(row=r, column=4).number_format = '$#,##0.00'
        for c in range(1, 6):
            ws6.cell(row=r, column=c).font = data_font
            ws6.cell(row=r, column=c).border = thin_border

    # Auto-adjust column widths for all sheets
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and ('$' in cell.number_format or '€' in cell.number_format):
                    val_str = f"${val_str}"  # Pad for formatting characters
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save to memory and return
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        download_name="Finance_Report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@export_bp.route('/export/tableau')
def export_tableau():
    user_id = session['user_id']
    
    # Columns: Date, Income, Expense, Category, Budget, Savings, Balance, Month, Year
    incomes = Income.query.filter_by(user_id=user_id).order_by(Income.date.asc()).all()
    expenses = Expense.query.filter_by(user_id=user_id).order_by(Expense.date.asc()).all()
    
    # Calculate running balance and compile list
    txs = []
    for i in incomes:
        txs.append({'date': i.date, 'income': i.amount, 'expense': 0.0, 'category': i.source, 'type': 'income'})
    for e in expenses:
        txs.append({'date': e.date, 'income': 0.0, 'expense': e.amount, 'category': e.category, 'type': 'expense'})
        
    txs.sort(key=lambda x: x['date'])
    
    running_balance = 0.0
    data = []
    
    for t in txs:
        running_balance += t['income'] - t['expense']
        month_str = t['date'].strftime('%B')
        year_val = t['date'].year
        month_num_str = t['date'].strftime('%Y-%m')
        
        # Get category budget
        budget_limit = 0.0
        if t['type'] == 'expense':
            b = Budget.query.filter(
                Budget.category == t['category'],
                Budget.month == month_num_str,
                Budget.user_id == user_id
            ).first()
            if b:
                budget_limit = b.limit_amount
                
        # Total Savings Goals Accumulated
        total_savings = db.session.query(db.func.sum(SavingsGoal.current_amount))\
            .filter(SavingsGoal.user_id == user_id).scalar() or 0.0
        
        data.append({
            'Date': t['date'].strftime('%Y-%m-%d'),
            'Income': t['income'],
            'Expense': t['expense'],
            'Category': t['category'],
            'Budget': budget_limit,
            'Savings': total_savings,
            'Balance': running_balance,
            'Month': month_str,
            'Year': year_val
        })
        
    if not data:
        # Avoid empty DataFrame error, create a dummy structure
        df = pd.DataFrame(columns=['Date', 'Income', 'Expense', 'Category', 'Budget', 'Savings', 'Balance', 'Month', 'Year'])
    else:
        df = pd.DataFrame(data)
    
    # Return as CSV
    csv_buf = io.StringIO()
    df.to_csv(csv_buf, index=False)
    
    return send_file(
        io.BytesIO(csv_buf.getvalue().encode('utf-8')),
        download_name="tableau_dataset.csv",
        as_attachment=True,
        mimetype="text/csv"
    )

@export_bp.route('/export/powerbi')
def export_powerbi():
    user_id = session['user_id']
    
    # Columns: Transaction ID, Date, Month, Year, Income, Expense, Category, Budget, Savings, Balance, Financial Score
    incomes = Income.query.filter_by(user_id=user_id).order_by(Income.date.asc()).all()
    expenses = Expense.query.filter_by(user_id=user_id).order_by(Expense.date.asc()).all()
    
    txs = []
    for i in incomes:
        txs.append({'db_id': f"INC-{i.id}", 'date': i.date, 'income': i.amount, 'expense': 0.0, 'category': i.source, 'type': 'income'})
    for e in expenses:
        txs.append({'db_id': f"EXP-{e.id}", 'date': e.date, 'income': 0.0, 'expense': e.amount, 'category': e.category, 'type': 'expense'})
        
    txs.sort(key=lambda x: x['date'])
    
    running_balance = 0.0
    data = []
    
    for idx, t in enumerate(txs, start=1):
        running_balance += t['income'] - t['expense']
        month_str = t['date'].strftime('%B')
        year_val = t['date'].year
        month_num_str = t['date'].strftime('%Y-%m')
        
        # Get category budget
        budget_limit = 0.0
        if t['type'] == 'expense':
            b = Budget.query.filter(
                Budget.category == t['category'],
                Budget.month == month_num_str,
                Budget.user_id == user_id
            ).first()
            if b:
                budget_limit = b.limit_amount
                
        # Total Savings Goal Accumulated
        total_savings = db.session.query(db.func.sum(SavingsGoal.current_amount))\
            .filter(SavingsGoal.user_id == user_id).scalar() or 0.0
        
        # Calculate score for that month/year
        score = calculate_historical_score(user_id, year_val, t['date'].month)
        
        data.append({
            'Transaction ID': f"TX-{idx:05d}",
            'Date': t['date'].strftime('%Y-%m-%d'),
            'Month': month_str,
            'Year': year_val,
            'Income': t['income'],
            'Expense': t['expense'],
            'Category': t['category'],
            'Budget': budget_limit,
            'Savings': total_savings,
            'Balance': running_balance,
            'Financial Score': score
        })
        
    if not data:
        df = pd.DataFrame(columns=['Transaction ID', 'Date', 'Month', 'Year', 'Income', 'Expense', 'Category', 'Budget', 'Savings', 'Balance', 'Financial Score'])
    else:
        df = pd.DataFrame(data)
    
    csv_buf = io.StringIO()
    df.to_csv(csv_buf, index=False)
    
    return send_file(
        io.BytesIO(csv_buf.getvalue().encode('utf-8')),
        download_name="powerbi_dataset.csv",
        as_attachment=True,
        mimetype="text/csv"
    )

# --- File Import / Upload Module ---
@export_bp.route('/import', methods=['GET', 'POST'])
def import_data():
    settings = get_current_settings()
    user_id = session['user_id']
    
    if request.method == 'POST':
        file = request.files.get('file')
        import_type = request.form.get('type') # 'income', 'expense', or 'unified'
        
        if not file or file.filename == '':
            flash("No file selected for import.", "danger")
            return redirect(url_for('export.import_data'))
            
        filename = file.filename
        file_ext = os.path.splitext(filename)[1].lower()
        
        try:
            # Parse Excel or CSV
            if file_ext == '.csv':
                df = pd.read_csv(io.BytesIO(file.read()))
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(io.BytesIO(file.read()))
            else:
                flash("Unsupported file format. Please upload a CSV or Excel (.xlsx) file.", "danger")
                return redirect(url_for('export.import_data'))
                
            # Create a case-insensitive column mapping
            col_map = {}
            for col in df.columns:
                cleaned = str(col).strip().lower()
                col_map[cleaned] = col
                
            # Detect dates
            date_key = None
            for candidate in ['date', 'transaction date', 'trans_date', 'date_str', 'timestamp']:
                if candidate in col_map:
                    date_key = col_map[candidate]
                    break
                    
            if not date_key:
                flash("Missing date column. Please ensure your file has a date column.", "danger")
                return redirect(url_for('export.import_data'))
                
            # Detect description
            desc_key = None
            for candidate in ['description', 'desc', 'notes', 'memo', 'details']:
                if candidate in col_map:
                    desc_key = col_map[candidate]
                    break
                    
            # Detect category/source
            cat_key = None
            for candidate in ['category', 'source', 'category/source', 'type', 'category_or_source']:
                if candidate in col_map:
                    cat_key = col_map[candidate]
                    break

            # Check if unified mode
            is_unified = (import_type == 'unified') or ('income' in col_map and 'expense' in col_map)
            
            success_count = 0
            fail_count = 0
            
            if is_unified:
                # Unified import expects 'income' and 'expense' columns
                income_col = col_map.get('income')
                expense_col = col_map.get('expense')
                
                if not income_col and not expense_col:
                    flash("Unified import requires 'Income' and/or 'Expense' columns.", "danger")
                    return redirect(url_for('export.import_data'))
                    
                for idx, row in df.iterrows():
                    try:
                        # Parse date
                        raw_date = row[date_key]
                        if isinstance(raw_date, pd.Timestamp):
                            parsed_date = raw_date.date()
                        else:
                            parsed_date = pd.to_datetime(raw_date).date()
                            
                        desc = str(row[desc_key]) if desc_key and not pd.isna(row[desc_key]) else ''
                        cat_or_source = str(row[cat_key]).strip() if cat_key and not pd.isna(row[cat_key]) else 'Other'
                        
                        added_any = False
                        
                        # Process income
                        if income_col and not pd.isna(row[income_col]):
                            inc_amt = float(row[income_col])
                            if inc_amt > 0:
                                new_inc = Income(date=parsed_date, source=cat_or_source, amount=inc_amt, description=desc, user_id=user_id)
                                db.session.add(new_inc)
                                added_any = True
                                
                        # Process expense
                        if expense_col and not pd.isna(row[expense_col]):
                            exp_amt = float(row[expense_col])
                            if exp_amt > 0:
                                new_exp = Expense(date=parsed_date, category=cat_or_source, amount=exp_amt, description=desc, user_id=user_id)
                                db.session.add(new_exp)
                                added_any = True
                                
                        if added_any:
                            success_count += 1
                        else:
                            fail_count += 1
                    except Exception:
                        fail_count += 1
            else:
                # Standard import (single type)
                # Find amount column
                amount_col = None
                amt_candidates = ['amount', 'value', 'val']
                if import_type == 'income':
                    amt_candidates.append('income')
                else:
                    amt_candidates.append('expense')
                    
                for candidate in amt_candidates:
                    if candidate in col_map:
                        amount_col = col_map[candidate]
                        break
                        
                if not amount_col:
                    flash(f"Amount column not found. Please ensure the file has a column like 'Amount' or '{'Income' if import_type == 'income' else 'Expense'}'.", "danger")
                    return redirect(url_for('export.import_data'))
                    
                # Find category/source
                cat_col = col_map.get('category') or col_map.get('source') or col_map.get('category/source') or col_map.get('type')
                
                # Check required columns
                if not cat_col:
                    flash(f"Category or Source column not found. Please ensure your file contains a category/source column.", "danger")
                    return redirect(url_for('export.import_data'))
                    
                for idx, row in df.iterrows():
                    try:
                        # Parse date
                        raw_date = row[date_key]
                        if isinstance(raw_date, pd.Timestamp):
                            parsed_date = raw_date.date()
                        else:
                            parsed_date = pd.to_datetime(raw_date).date()
                            
                        # Parse amount
                        amount = float(row[amount_col])
                        if pd.isna(amount) or amount < 0:
                            fail_count += 1
                            continue
                            
                        desc = str(row[desc_key]) if desc_key and not pd.isna(row[desc_key]) else ''
                        cat_or_source = str(row[cat_col]).strip()
                        
                        if import_type == 'income':
                            new_row = Income(date=parsed_date, source=cat_or_source, amount=amount, description=desc, user_id=user_id)
                        else:
                            new_row = Expense(date=parsed_date, category=cat_or_source, amount=amount, description=desc, user_id=user_id)
                            
                        db.session.add(new_row)
                        success_count += 1
                    except Exception:
                        fail_count += 1
                        
            db.session.commit()
            flash(f"Import finished. Successfully imported {success_count} records. Failed records: {fail_count}.", "success" if fail_count == 0 else "warning")
            return redirect(url_for('dashboard.home'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error parsing file: {str(e)}", "danger")
            return redirect(url_for('export.import_data'))
            
    return render_template('import.html', settings=settings)
